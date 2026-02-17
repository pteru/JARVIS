#!/usr/bin/env python3
"""JARVIS XLSX Tool — Create, read, edit, and inspect Excel files."""

import argparse
import json
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import coordinate_from_string


def color_to_hex(color):
    """Convert an openpyxl Color object to #RRGGBB or None."""
    if color is None:
        return None
    if color.type == "rgb" and color.rgb and color.rgb != "00000000":
        rgb = str(color.rgb)
        if len(rgb) == 8:
            return f"#{rgb[2:]}"
        if len(rgb) == 6:
            return f"#{rgb}"
    if color.type == "theme":
        return f"theme:{color.theme}"
    if color.type == "indexed" and color.indexed is not None:
        return f"indexed:{color.indexed}"
    return None


def border_to_dict(border):
    """Convert Border to a simple dict."""
    if border is None:
        return None
    sides = {}
    for name in ("left", "right", "top", "bottom"):
        side = getattr(border, name, None)
        if side and side.style:
            sides[name] = side.style
    return sides if sides else None


def cell_style_dict(cell):
    """Extract style info from a cell."""
    style = {}
    if cell.fill and cell.fill.fgColor:
        c = color_to_hex(cell.fill.fgColor)
        if c:
            style["fill"] = c
    if cell.font:
        fc = color_to_hex(cell.font.color)
        if fc:
            style["font_color"] = fc
        if cell.font.bold:
            style["bold"] = True
        if cell.font.italic:
            style["italic"] = True
    b = border_to_dict(cell.border)
    if b:
        style["borders"] = b
    return style if style else None


def parse_range(ws, range_str):
    """Parse a range string like A1:D10 and return (min_row, min_col, max_row, max_col)."""
    if ":" in range_str:
        start, end = range_str.split(":")
        sc, sr = coordinate_from_string(start)
        ec, er = coordinate_from_string(end)
        return sr, column_index_from_string(sc), er, column_index_from_string(ec)
    else:
        sc, sr = coordinate_from_string(range_str)
        ci = column_index_from_string(sc)
        return sr, ci, sr, ci


def cmd_read(args):
    wb = load_workbook(args.file, data_only=True)
    ws = wb[args.sheet] if args.sheet else wb.active

    if args.range:
        min_r, min_c, max_r, max_c = parse_range(ws, args.range)
    else:
        min_r, min_c = ws.min_row, ws.min_column
        max_r, max_c = ws.max_row, ws.max_column

    rows = []
    for r in range(min_r, max_r + 1):
        row_data = []
        for c in range(min_c, max_c + 1):
            cell = ws.cell(row=r, column=c)
            val = cell.value
            entry = val
            if args.with_styles:
                st = cell_style_dict(cell)
                entry = {"value": val}
                if st:
                    entry["style"] = st
            row_data.append(entry)
        rows.append(row_data)

    result = {"sheet": ws.title, "rows": rows}
    if args.with_styles and ws.merged_cells.ranges:
        result["merged"] = [str(m) for m in ws.merged_cells.ranges]

    if args.format == "json":
        print(json.dumps(result, default=str, indent=2))
    else:
        for row in rows:
            vals = []
            for v in row:
                if isinstance(v, dict):
                    v = v.get("value")
                vals.append(str(v) if v is not None else "")
            print("\t".join(vals))


def cmd_create(args):
    if args.input and args.input != "-":
        data = json.loads(Path(args.input).read_text())
    else:
        data = json.loads(sys.stdin.read())

    wb = Workbook()
    ws = wb.active
    if args.sheet:
        ws.title = args.sheet

    if isinstance(data, list) and len(data) > 0:
        if isinstance(data[0], dict):
            headers = list(data[0].keys())
            for c, h in enumerate(headers, 1):
                ws.cell(row=1, column=c, value=h)
            for r, record in enumerate(data, 2):
                for c, h in enumerate(headers, 1):
                    ws.cell(row=r, column=c, value=record.get(h))
        elif isinstance(data[0], list):
            for r, row in enumerate(data, 1):
                for c, val in enumerate(row, 1):
                    ws.cell(row=r, column=c, value=val)

    wb.save(args.file)
    print(f"Created {args.file}")


def cmd_edit(args):
    wb = load_workbook(args.file)
    ws = wb[args.sheet] if args.sheet else wb.active

    if args.set:
        for assignment in args.set:
            cell_ref, value = assignment.split("=", 1)
            # Try numeric conversion
            try:
                value = int(value)
            except ValueError:
                try:
                    value = float(value)
                except ValueError:
                    pass
            ws[cell_ref.strip()] = value

    if args.insert_row:
        ws.insert_rows(args.insert_row)

    if args.delete_row:
        ws.delete_rows(args.delete_row)

    if args.rename:
        ws.title = args.rename

    wb.save(args.file)
    print(f"Edited {args.file}")


def cmd_info(args):
    wb = load_workbook(args.file, data_only=True)
    info = {"file": args.file, "sheets": []}
    for name in wb.sheetnames:
        ws = wb[name]
        sheet_info = {
            "name": name,
            "dimensions": ws.dimensions,
        }
        # Read first row as headers
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1), []):
            headers.append(str(cell.value) if cell.value is not None else "")
        if any(headers):
            sheet_info["headers"] = headers
        info["sheets"].append(sheet_info)
    wb.close()
    print(json.dumps(info, indent=2))


def main():
    parser = argparse.ArgumentParser(description="JARVIS XLSX Tool")
    sub = parser.add_subparsers(dest="command", required=True)

    # read
    p = sub.add_parser("read", help="Read XLSX data")
    p.add_argument("file")
    p.add_argument("--sheet")
    p.add_argument("--range", help="Cell range e.g. A1:D10")
    p.add_argument("--format", choices=["json", "table"], default="table")
    p.add_argument("--with-styles", action="store_true")

    # create
    p = sub.add_parser("create", help="Create XLSX from JSON")
    p.add_argument("file")
    p.add_argument("--input", "-i", help="JSON file (default: stdin)")
    p.add_argument("--sheet", help="Sheet name")

    # edit
    p = sub.add_parser("edit", help="Edit XLSX file")
    p.add_argument("file")
    p.add_argument("--sheet")
    p.add_argument("--set", action="append", help="Set cell e.g. A1=Hello")
    p.add_argument("--insert-row", type=int)
    p.add_argument("--delete-row", type=int)
    p.add_argument("--rename", help="Rename sheet")

    # info
    p = sub.add_parser("info", help="Show XLSX info")
    p.add_argument("file")

    args = parser.parse_args()
    {"read": cmd_read, "create": cmd_create, "edit": cmd_edit, "info": cmd_info}[args.command](args)


if __name__ == "__main__":
    main()
